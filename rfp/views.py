# rfp/views.py

from decimal import Decimal, InvalidOperation
from datetime import timedelta
import random
import logging
import traceback

from openpyxl import Workbook

from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Q, Count
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods, require_POST
from django.utils import timezone

from .models import Vendor, Category, RFP, Quote, QuoteItem, AuthConfig, LoginOTP

logger = logging.getLogger(__name__)


# ---------------- AUTH CONFIG HELPER ----------------
def get_auth_config():
    """
    Creates config row if not present.
    """
    cfg, _ = AuthConfig.objects.get_or_create(
        id=1,
        defaults={
            "enable_vendor_2fa": True,
            "otp_expiry_minutes": 10,
        }
    )
    return cfg


# ---------------- LOGIN ----------------
def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser:
            return redirect("admin_dashboard")
        return redirect("vendor_dashboard")

    if request.method == "POST":
        email = (request.POST.get("email") or "").strip().lower()
        password = request.POST.get("password") or ""

        user = authenticate(request, username=email, password=password)
        if not user:
            return render(request, "rfp/login.html", {"error": "Invalid credentials"})

        # Admin login directly (no OTP)
        if user.is_staff or user.is_superuser:
            login(request, user)
            return redirect("admin_dashboard")

        # Vendor must exist
        vendor = Vendor.objects.filter(email=email).first()
        if not vendor:
            return render(request, "rfp/login.html", {"error": "Vendor profile not found. Contact admin."})

        # Vendor must be APPROVED
        if vendor.status != Vendor.ApprovalStatus.APPROVED:
            return render(request, "rfp/login.html", {"error": "Your account is not approved yet."})

        # Check backend config
        cfg = get_auth_config()
        if not cfg.enable_vendor_2fa:
            login(request, user)
            return redirect("vendor_dashboard")

        # Invalidate previous OTPs
        LoginOTP.objects.filter(email=email, is_used=False).update(is_used=True)

        # Create OTP + expiry
        otp = LoginOTP.generate_otp()
        expires_at = timezone.now() + timedelta(minutes=int(cfg.otp_expiry_minutes or 10))

        # Store pending login in session FIRST
        request.session["pending_2fa_email"] = email
        request.session["pending_2fa_user_id"] = user.id

        # Create OTP row + Send OTP (email) safely (no worker-kill mystery)
        otp_row = None
        try:
            otp_row = LoginOTP.objects.create(
                email=email,
                otp=otp,
                expires_at=expires_at
            )

            send_mail(
                subject="Your Login OTP",
                message=f"Your OTP is: {otp}. It will expire in {cfg.otp_expiry_minutes} minutes.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )

        except Exception as e:
            # Log full error to Render logs
            logger.exception("OTP email send failed for %s", email)
            traceback.print_exc()

            # cleanup session
            request.session.pop("pending_2fa_email", None)
            request.session.pop("pending_2fa_user_id", None)

            # delete otp row if created
            if otp_row:
                otp_row.delete()

            return render(request, "rfp/login.html", {
                "error": f"OTP email failed: {str(e)}"
            })

        return redirect("vendor_verify_otp")

    return render(request, "rfp/login.html")


# ---------------- VERIFY OTP ----------------
@require_http_methods(["GET", "POST"])
def vendor_verify_otp(request):
    email = request.session.get("pending_2fa_email")
    user_id = request.session.get("pending_2fa_user_id")

    if not email or not user_id:
        return redirect("login")

    if request.method == "POST":
        otp_entered = (request.POST.get("otp") or "").strip()

        otp_obj = (
            LoginOTP.objects
            .filter(email=email, otp=otp_entered, is_used=False)
            .order_by("-created_at")
            .first()
        )

        if not otp_obj:
            return render(request, "rfp/vendor_verify_otp.html", {"error": "Invalid OTP."})

        if timezone.now() > otp_obj.expires_at:
            return render(request, "rfp/vendor_verify_otp.html", {"error": "OTP expired. Please resend OTP."})

        # mark used
        otp_obj.is_used = True
        otp_obj.save(update_fields=["is_used"])

        # login user
        user = User.objects.filter(id=user_id).first()
        if not user:
            return redirect("login")

        login(request, user)

        # clear session
        request.session.pop("pending_2fa_email", None)
        request.session.pop("pending_2fa_user_id", None)

        return redirect("vendor_dashboard")

    return render(request, "rfp/vendor_verify_otp.html")


# ---------------- RESEND OTP ----------------
@require_POST
def vendor_resend_otp(request):
    email = request.session.get("pending_2fa_email")
    user_id = request.session.get("pending_2fa_user_id")

    if not email or not user_id:
        return redirect("login")

    cfg = get_auth_config()

    # invalidate old OTPs
    LoginOTP.objects.filter(email=email, is_used=False).update(is_used=True)

    otp = LoginOTP.generate_otp()
    expires_at = timezone.now() + timedelta(minutes=int(cfg.otp_expiry_minutes or 10))

    otp_row = None
    try:
        otp_row = LoginOTP.objects.create(email=email, otp=otp, expires_at=expires_at)

        send_mail(
            subject="Your Login OTP (Resent)",
            message=f"Your OTP is: {otp}. It will expire in {cfg.otp_expiry_minutes} minutes.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False
        )

    except Exception as e:
        logger.exception("Resend OTP email failed for %s", email)
        traceback.print_exc()

        if otp_row:
            otp_row.delete()

        return render(request, "rfp/vendor_verify_otp.html", {"error": f"OTP resend failed: {e}"})

    return redirect("vendor_verify_otp")


# ---------------- LOGOUT ----------------
@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


# ---------------- SIGNUP ----------------
def admin_signup(request):
    error = None

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = (request.POST.get("email") or "").strip().lower()
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            error = "Passwords do not match."
        else:
            try:
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                user.is_staff = True
                user.save()
                return redirect("login")
            except IntegrityError:
                error = "Email already exists."

    return render(request, "rfp/admin_signup.html", {"error": error})


def vendor_signup(request):
    categories = Category.objects.filter(status=Category.Status.ACTIVE)

    if request.method == "POST":
        f_name = (request.POST.get("f_name") or "").strip()
        l_name = (request.POST.get("l_name") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        phone = (request.POST.get("phone") or "").strip()
        password = request.POST.get("password") or ""
        confirm_password = request.POST.get("confirm_password") or ""

        revenue_raw = (request.POST.get("revenue") or "").strip()
        employees_raw = (request.POST.get("employees") or "").strip()
        gst_no = (request.POST.get("gst") or "").strip()
        pan_no = (request.POST.get("pan") or "").strip()
        category_id = (request.POST.get("category_id") or "").strip()

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, "rfp/vendor_signup.html", {"categories": categories})

        if User.objects.filter(username=email).exists():
            messages.error(request, "Email already exists.")
            return render(request, "rfp/vendor_signup.html", {"categories": categories})

        if Vendor.objects.filter(contact=phone).exists():
            messages.error(request, "Phone already exists.")
            return render(request, "rfp/vendor_signup.html", {"categories": categories})

        revenue = None
        try:
            revenue = Decimal(revenue_raw) if revenue_raw else None
        except (InvalidOperation, ValueError):
            revenue = None

        employees = None
        try:
            employees = int(employees_raw) if employees_raw else None
        except ValueError:
            employees = None

        category = Category.objects.filter(
            id=category_id,
            status=Category.Status.ACTIVE
        ).first()

        if not category:
            messages.error(request, "Please select a valid category.")
            return render(request, "rfp/vendor_signup.html", {"categories": categories})

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password
                )

                Vendor.objects.create(
                    first_name=f_name,
                    last_name=l_name,
                    email=email,
                    contact=phone,
                    revenue_last_3_years_lakhs=revenue,
                    employees_count=employees,
                    gst_no=gst_no,
                    pan_no=pan_no,
                    category=category,
                    status=Vendor.ApprovalStatus.PENDING,
                )

            login(request, user)
            messages.success(request, "Registered successfully! Wait for admin approval.")
            return redirect("vendor_dashboard")

        except IntegrityError:
            messages.error(request, "Something went wrong. Email/phone may already exist.")
            return render(request, "rfp/vendor_signup.html", {"categories": categories})

    return render(request, "rfp/vendor_signup.html", {"categories": categories})


# ---------------- PASSWORD RESET ----------------
def forgot_password(request):
    message = None
    error = None
    otp_sent = request.session.get("otp_sent", False)
    email = request.session.get("reset_email", "")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "send_otp":
            email = (request.POST.get("email") or "").strip().lower()

            if not User.objects.filter(email=email).exists():
                error = "Email not registered."
            else:
                otp = str(random.randint(100000, 999999))
                request.session["reset_email"] = email
                request.session["reset_otp"] = otp
                request.session["otp_sent"] = True

                send_mail(
                    subject="Your OTP for Password Reset",
                    message=f"Your OTP is: {otp}",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                message = "OTP sent to your registered email."
                otp_sent = True

        elif action == "verify_otp":
            otp_entered = (request.POST.get("otp") or "").strip()
            otp_saved = request.session.get("reset_otp")

            if otp_saved and otp_entered == otp_saved:
                request.session["otp_verified"] = True
                return redirect("reset_password")
            else:
                error = "Invalid OTP. Please try again."
                otp_sent = True

    return render(request, "rfp/forgot_password.html", {
        "message": message,
        "error": error,
        "otp_sent": otp_sent,
        "email": email
    })


def reset_password(request):
    if not request.session.get("otp_verified"):
        return redirect("forgot_password")

    email = request.session.get("reset_email")
    error = None

    if request.method == "POST":
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            error = "Passwords do not match."
        else:
            user = User.objects.filter(email=email).first()
            if not user:
                error = "User not found."
            else:
                user.set_password(password)
                user.save()

                for k in ["reset_email", "reset_otp", "otp_sent", "otp_verified"]:
                    request.session.pop(k, None)

                messages.success(request, "Password reset successful. Please login.")
                return redirect("login")

    return render(request, "rfp/reset_password.html", {"error": error})


# ---------------- ADMIN HELPERS ----------------
def _admin_only(request):
    return request.user.is_authenticated and request.user.is_staff


# ---------------- ADMIN PANEL ----------------
@login_required
def admin_dashboard(request):
    if not _admin_only(request):
        return redirect("login")
    return render(request, "rfp/admin_dashboard.html")


@login_required
def admin_categories(request):
    if not _admin_only(request):
        return redirect("login")

    q = (request.GET.get("q") or "").strip()

    qs = Category.objects.all().order_by("id")
    if q:
        qs = qs.filter(name__icontains=q)

    page_obj = Paginator(qs, 10).get_page(request.GET.get("page"))

    return render(request, "rfp/admin_categories.html", {"page_obj": page_obj, "q": q})


@login_required
def admin_category_create(request):
    if not _admin_only(request):
        return redirect("login")

    error = None

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()

        if not name:
            error = "Category name is required."
        elif Category.objects.filter(name__iexact=name).exists():
            error = "Category already exists."
        else:
            Category.objects.create(name=name, status=Category.Status.ACTIVE)
            messages.success(request, "Category added successfully.")
            return redirect("admin_categories")

    return render(request, "rfp/admin_category_create.html", {"error": error})


@login_required
def admin_category_toggle(request, pk):
    if not _admin_only(request):
        return redirect("login")

    cat = Category.objects.filter(pk=pk).first()
    if not cat:
        messages.error(request, "Category not found.")
        return redirect("admin_categories")

    cat.status = Category.Status.INACTIVE if cat.status == Category.Status.ACTIVE else Category.Status.ACTIVE
    cat.save(update_fields=["status"])

    messages.success(request, "Status updated.")
    return redirect("admin_categories")


@login_required
def admin_vendors(request):
    if not _admin_only(request):
        return redirect("login")

    qs = Vendor.objects.all().order_by("-id")
    page_obj = Paginator(qs, 10).get_page(request.GET.get("page"))

    return render(request, "rfp/admin_vendors.html", {"page_obj": page_obj})


@login_required
def admin_vendor_toggle(request, pk):
    if not _admin_only(request):
        return redirect("login")

    vendor = Vendor.objects.filter(pk=pk).first()
    if not vendor:
        messages.error(request, "Vendor not found.")
        return redirect("admin_vendors")

    vendor.status = (
        Vendor.ApprovalStatus.REJECTED
        if vendor.status == Vendor.ApprovalStatus.APPROVED
        else Vendor.ApprovalStatus.APPROVED
    )
    vendor.save(update_fields=["status"])

    messages.success(request, "Vendor status updated.")
    return redirect("admin_vendors")


@login_required
def admin_rfp_list(request):
    if not _admin_only(request):
        return redirect("login")

    qs = RFP.objects.select_related("category").order_by("-id")
    page_obj = Paginator(qs, 10).get_page(request.GET.get("page"))

    return render(request, "rfp/admin_rfp_list.html", {"page_obj": page_obj})


@login_required
def admin_rfp_select_category(request):
    if not request.user.is_staff:
        return redirect("login")

    categories = Category.objects.filter(status=Category.Status.ACTIVE)

    if request.method == "POST":
        category_id = request.POST.get("category")

        if not category_id:
            messages.error(request, "Please select category.")
            return redirect("admin_rfp_select_category")

        return redirect("admin_rfp_add", category_id=category_id)

    return render(request, "rfp/admin_rfp_select_category.html", {"categories": categories})


@login_required
def admin_rfp_add(request, category_id):
    if not request.user.is_staff:
        return redirect("login")

    category = Category.objects.filter(id=category_id, status=Category.Status.ACTIVE).first()
    if not category:
        messages.error(request, "Invalid category.")
        return redirect("admin_rfp_select_category")

    vendors = Vendor.objects.filter(
        status=Vendor.ApprovalStatus.APPROVED,
        category=category
    ).order_by("-id")

    if request.method == "POST":
        title = request.POST.get("title")
        last_date = request.POST.get("last_date")
        min_amount = request.POST.get("min_amount")
        max_amount = request.POST.get("max_amount")
        vendor_ids = request.POST.getlist("vendors")

        if not title or not last_date or not min_amount or not max_amount:
            messages.error(request, "All fields are required.")
            return redirect("admin_rfp_add", category_id=category.id)

        rfp = RFP.objects.create(
            category=category,
            title=title,
            last_date=last_date,
            min_amount=min_amount,
            max_amount=max_amount,
            status=RFP.Status.OPEN
        )
        rfp.assigned_vendors.set(vendor_ids)

        messages.success(request, "RFP created successfully.")
        return redirect("admin_rfp_list")

    return render(request, "rfp/admin_rfp_add.html", {"category": category, "vendors": vendors})


@login_required
def admin_rfp_toggle(request, pk):
    if not _admin_only(request):
        return redirect("login")

    obj = RFP.objects.filter(pk=pk).first()
    if not obj:
        messages.error(request, "RFP not found.")
        return redirect("admin_rfp_list")

    obj.status = RFP.Status.CLOSED if obj.status == RFP.Status.OPEN else RFP.Status.OPEN
    obj.save(update_fields=["status"])

    messages.success(request, "RFP status updated.")
    return redirect("admin_rfp_list")


@login_required
def admin_quotes(request):
    if not _admin_only(request):
        return redirect("login")

    qs = QuoteItem.objects.select_related(
        "quote", "quote__rfp", "quote__vendor"
    ).order_by("-id")

    page_obj = Paginator(qs, 10).get_page(request.GET.get("page"))

    return render(request, "rfp/admin_quotes.html", {"page_obj": page_obj})


# ---------------- VENDOR HELPERS ----------------
def _get_vendor(request):
    email = (request.user.email or request.user.username or "").strip().lower()
    return Vendor.objects.filter(email=email).first()


def _vendor_only(request):
    if not request.user.is_authenticated:
        return False
    if request.user.is_staff or request.user.is_superuser:
        return False
    vendor = _get_vendor(request)
    return vendor and vendor.status == Vendor.ApprovalStatus.APPROVED


# ---------------- VENDOR PANEL ----------------
@login_required
def vendor_dashboard(request):
    if request.user.is_staff:
        return redirect("admin_dashboard")

    email = request.user.username
    vendor = Vendor.objects.filter(email=email).first()
    if not vendor:
        return redirect("login")

    return render(request, "rfp/vendor_dashboard.html", {"vendor": vendor})


@login_required
def vendor_rfp_list(request):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    rfps = RFP.objects.select_related("category").filter(
        status=RFP.Status.OPEN,
        assigned_vendors=vendor
    ).order_by("-id")

    return render(request, "rfp/rfp_list.html", {"rfps": rfps})


@login_required
def vendor_rfp_detail(request, pk):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    rfp = get_object_or_404(RFP.objects.select_related("category"), pk=pk)

    if rfp.status != RFP.Status.OPEN:
        messages.error(request, "This RFP is closed.")
        return redirect("vendor_rfp_list")

    existing_quote = Quote.objects.filter(rfp=rfp, vendor=vendor).first()

    return render(request, "rfp/rfp_detail.html", {"rfp": rfp, "existing_quote": existing_quote})


@login_required
def vendor_quote_list(request):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    quotes = vendor.quotes.select_related("rfp", "rfp__category").order_by("-id")
    return render(request, "rfp/quote_list.html", {"quotes": quotes})


@login_required
@require_http_methods(["GET", "POST"])
def vendor_quote_create(request, rfp_id):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    rfp = get_object_or_404(RFP.objects.select_related("category"), pk=rfp_id)

    existing = Quote.objects.filter(rfp=rfp, vendor=vendor).first()
    if existing:
        return redirect("vendor_quote_submit", rfp_id=rfp.id)

    if request.method == "POST":
        Quote.objects.create(rfp=rfp, vendor=vendor)
        messages.success(request, "Quote created. Add items now.")
        return redirect("vendor_quote_submit", rfp_id=rfp.id)

    return render(request, "rfp/quote_create.html", {"rfp": rfp})


@login_required
@require_http_methods(["GET", "POST"])
def vendor_quote_submit(request, rfp_id):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    rfp = get_object_or_404(RFP.objects.select_related("category"), pk=rfp_id)

    quote = Quote.objects.filter(rfp=rfp, vendor=vendor).first()
    if not quote:
        return redirect("vendor_quote_create", rfp_id=rfp.id)

    if request.method == "POST":
        item_names = request.POST.getlist("item_name")
        prices = request.POST.getlist("vendor_price")
        quantities = request.POST.getlist("quantity")

        if not item_names:
            messages.error(request, "Please add at least one item.")
            return redirect("vendor_quote_submit", rfp_id=rfp.id)

        with transaction.atomic():
            quote.items.all().delete()

            for i in range(len(item_names)):
                name = (item_names[i] or "").strip()
                if not name:
                    continue

                raw_price = (prices[i] if i < len(prices) else "") or "0"
                try:
                    price = Decimal(str(raw_price).strip() or "0")
                except (InvalidOperation, TypeError, ValueError):
                    price = Decimal("0")

                raw_qty = (quantities[i] if i < len(quantities) else "") or "1"
                try:
                    qty = int(str(raw_qty).strip() or "1")
                except (TypeError, ValueError):
                    qty = 1

                QuoteItem.objects.create(
                    quote=quote,
                    item_name=name,
                    vendor_price=price,
                    quantity=qty
                )

        messages.success(request, "Quote submitted successfully.")
        return redirect("vendor_quote_list")

    return render(request, "rfp/quote_submit.html", {"rfp": rfp, "quote": quote, "items": quote.items.all()})


@login_required
def vendor_rfp_for_quotes(request):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    qs = RFP.objects.select_related("category").filter(
        status=RFP.Status.OPEN,
        assigned_vendors=vendor
    ).order_by("-id")

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    rfp_ids = [r.id for r in page_obj.object_list]
    quoted_ids = set(
        Quote.objects.filter(vendor=vendor, rfp_id__in=rfp_ids).values_list("rfp_id", flat=True)
    )

    return render(request, "rfp/rfp_for_quotes.html", {"page_obj": page_obj, "quoted_ids": quoted_ids})


@login_required
@require_http_methods(["GET", "POST"])
def vendor_apply_rfp(request, rfp_id):
    if not _vendor_only(request):
        return redirect("admin_dashboard")

    vendor = _get_vendor(request)
    if not vendor:
        messages.error(request, "Vendor profile not found.")
        return redirect("login")

    rfp = get_object_or_404(RFP.objects.select_related("category"), id=rfp_id)

    if rfp.status != RFP.Status.OPEN:
        messages.error(request, "This RFP is closed.")
        return redirect("vendor_rfp_for_quotes")

    quote, _ = Quote.objects.get_or_create(rfp=rfp, vendor=vendor)

    if request.method == "GET":
        return render(request, "rfp/quote_submit.html", {"rfp": rfp, "quote": quote, "items": quote.items.all()})

    item_names = request.POST.getlist("item_name")
    prices = request.POST.getlist("vendor_price")
    quantities = request.POST.getlist("quantity")

    if not item_names:
        messages.error(request, "Please add at least one item.")
        return redirect("vendor_apply_rfp", rfp_id=rfp.id)

    with transaction.atomic():
        quote.items.all().delete()

        for i in range(len(item_names)):
            name = (item_names[i] or "").strip()
            if not name:
                continue

            raw_price = (prices[i] if i < len(prices) else "") or "0"
            try:
                price = Decimal(str(raw_price).strip() or "0")
            except (InvalidOperation, TypeError, ValueError):
                price = Decimal("0")

            raw_qty = (quantities[i] if i < len(quantities) else "") or "1"
            try:
                qty = int(str(raw_qty).strip() or "1")
            except (TypeError, ValueError):
                qty = 1

            QuoteItem.objects.create(
                quote=quote,
                item_name=name,
                vendor_price=price,
                quantity=qty
            )

    messages.success(request, "Quote submitted successfully.")
    return redirect("vendor_quote_list")


# ---------------- EXPORT EXCEL ----------------
@login_required
def export_categories_excel(request):
    if not _admin_only(request):
        return redirect("login")

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"

    ws.append(["ID", "Category Name", "Status", "Created At"])
    categories = Category.objects.all().order_by("id")

    for c in categories:
        ws.append([c.id, c.name, c.status, c.created_at.strftime("%Y-%m-%d %H:%M")])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=categories.xlsx"
    wb.save(response)
    return response


@login_required
def export_vendors_excel(request):
    if not _admin_only(request):
        return redirect("login")

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"

    ws.append(["ID", "Name", "Email", "Phone", "Category", "Revenue (Lakhs)", "Employees", "GST", "PAN", "Status"])

    vendors = Vendor.objects.select_related("category").all().order_by("id")
    for v in vendors:
        ws.append([
            v.id,
            f"{v.first_name} {v.last_name}",
            v.email,
            v.contact,
            v.category.name if v.category else "",
            v.revenue_last_3_years_lakhs,
            v.employees_count,
            v.gst_no,
            v.pan_no,
            v.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=vendors.xlsx"
    wb.save(response)
    return response


@login_required
def export_rfp_excel(request):
    if not _admin_only(request):
        return redirect("login")

    wb = Workbook()
    ws = wb.active
    ws.title = "RFP List"

    ws.append(["ID", "Title", "Category", "Last Date", "Min Amount", "Max Amount", "Status"])

    rfps = RFP.objects.select_related("category").all().order_by("id")
    for r in rfps:
        ws.append([
            r.id,
            r.title,
            r.category.name,
            r.last_date.strftime("%Y-%m-%d"),
            r.min_amount,
            r.max_amount,
            r.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=rfp_list.xlsx"
    wb.save(response)
    return response


# ---------------- REPORTS ----------------
@login_required
def admin_reports(request):
    if not _admin_only(request):
        return redirect("login")

    total_vendors = Vendor.objects.count()
    approved_vendors = Vendor.objects.filter(status=Vendor.ApprovalStatus.APPROVED).count()
    pending_vendors = Vendor.objects.filter(status=Vendor.ApprovalStatus.PENDING).count()
    rejected_vendors = Vendor.objects.filter(status=Vendor.ApprovalStatus.REJECTED).count()

    vendor_category_stats = Category.objects.annotate(total=Count("vendors")).values("name", "total")

    total_rfp = RFP.objects.count()
    open_rfp = RFP.objects.filter(status=RFP.Status.OPEN).count()
    closed_rfp = RFP.objects.filter(status=RFP.Status.CLOSED).count()

    expired_rfp = RFP.objects.filter(
        status=RFP.Status.OPEN,
        last_date__lt=timezone.now().date()
    ).count()

    rfp_category_stats = Category.objects.annotate(total=Count("rfps")).values("name", "total")

    context = {
        "total_vendors": total_vendors,
        "approved_vendors": approved_vendors,
        "pending_vendors": pending_vendors,
        "rejected_vendors": rejected_vendors,
        "vendor_category_stats": vendor_category_stats,
        "total_rfp": total_rfp,
        "open_rfp": open_rfp,
        "closed_rfp": closed_rfp,
        "expired_rfp": expired_rfp,
        "rfp_category_stats": rfp_category_stats,
    }

    return render(request, "rfp/admin_reports.html", context)